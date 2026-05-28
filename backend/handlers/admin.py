"""
Задача 15 — /benchmark: сравнение Groq vs Ollama (20 вопросов, Excel-отчёт).
Задача 17 — /model pull <name>: загрузка модели Ollama с прогресс-баром.
"""
from __future__ import annotations

import asyncio
import io
import time
from typing import List

from aiogram import Router
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, Message

from config.settings import settings
from services.groq_service import groq_service, TaskType
from services.ollama_service import ollama_service
from utils.logging import get_logger

logger = get_logger(__name__)
router = Router(name="admin")

BENCHMARK_QUESTIONS = [
    "Что такое машинное обучение?",
    "Как работает нейронная сеть?",
    "Объясни принцип обратного распространения ошибки.",
    "Что такое трансформерная архитектура?",
    "Чем отличается supervised от unsupervised learning?",
    "Что такое переобучение и как его избежать?",
    "Опиши алгоритм градиентного спуска.",
    "Что такое attention mechanism?",
    "Как работает BERT?",
    "Что такое RAG (Retrieval-Augmented Generation)?",
    "Объясни концепцию embedding-векторов.",
    "Чем отличается GPT от BERT?",
    "Что такое квантизация модели?",
    "Как работает LoRA fine-tuning?",
    "Что такое prompt engineering?",
    "Объясни chain-of-thought prompting.",
    "Что такое temperature в языковых моделях?",
    "Как работает токенизация текста?",
    "Что такое контекстное окно модели?",
    "Объясни разницу между inference и training.",
]


@router.message(Command("benchmark"))
async def cmd_benchmark(message: Message) -> None:
    """Задача 15: запускает бенчмарк и выгружает Excel-отчёт."""
    sent = await message.answer("📊 Запускаю бенчмарк (20 вопросов)… Это займёт 1–2 минуты.")

    groq_results = []
    ollama_results = []

    ollama_available = await ollama_service.is_available()

    for i, question in enumerate(BENCHMARK_QUESTIONS, 1):
        await sent.edit_text(f"📊 Бенчмарк: {i}/20 вопросов…")

        # Groq
        t0 = time.monotonic()
        try:
            resp = await groq_service.chat(
                [
                    {"role": "system", "content": "Отвечай кратко, 2-3 предложения."},
                    {"role": "user", "content": question},
                ],
                task_type=TaskType.SHORT,
            )
            groq_latency = time.monotonic() - t0
            approx_tokens = len(resp.split()) * 1.3
            groq_results.append({
                "question": question,
                "latency_s": round(groq_latency, 3),
                "response_len": len(resp),
                "tokens_per_sec": round(approx_tokens / groq_latency, 1),
                "status": "ok",
            })
        except Exception as exc:
            groq_results.append({"question": question, "status": f"error: {exc}"})

        # Ollama
        if ollama_available:
            result = await ollama_service.benchmark_question(question)
            ollama_results.append(result)
        else:
            ollama_results.append({"question": question, "status": "unavailable"})

    # Build Excel report
    excel_bytes = _build_excel_report(groq_results, ollama_results)
    await sent.edit_text("✅ Бенчмарк завершён!")
    await message.answer_document(
        BufferedInputFile(excel_bytes, filename="benchmark_report.xlsx"),
        caption="📊 Отчёт Groq vs Ollama",
    )


def _build_excel_report(groq_results: list, ollama_results: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Groq sheet ──────────────────────────────────────────────────────
    ws_g = wb.active
    ws_g.title = "Groq"
    headers = ["#", "Вопрос", "Latency (s)", "Длина ответа", "Tokens/sec", "Статус"]
    _write_headers(ws_g, headers)
    for i, r in enumerate(groq_results, 1):
        ws_g.append([
            i,
            r.get("question", ""),
            r.get("latency_s", ""),
            r.get("response_len", ""),
            r.get("tokens_per_sec", ""),
            r.get("status", ""),
        ])

    # ── Ollama sheet ────────────────────────────────────────────────────
    ws_o = wb.create_sheet("Ollama")
    _write_headers(ws_o, headers)
    for i, r in enumerate(ollama_results, 1):
        ws_o.append([
            i,
            r.get("question", ""),
            r.get("latency_s", ""),
            r.get("response_len", ""),
            r.get("tokens_per_sec", ""),
            r.get("status", "ok"),
        ])

    # ── Summary sheet ───────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.append(["Метрика", "Groq", "Ollama"])
    _write_headers(ws_s, ["Метрика", "Groq", "Ollama"])

    def avg(data, key):
        vals = [r[key] for r in data if isinstance(r.get(key), (int, float))]
        return round(sum(vals) / len(vals), 3) if vals else "N/A"

    def percentile(data, key, p):
        vals = sorted(r[key] for r in data if isinstance(r.get(key), (int, float)))
        if not vals:
            return "N/A"
        idx = int(len(vals) * p / 100)
        return round(vals[min(idx, len(vals) - 1)], 3)

    metrics = [
        ("Avg latency (s)", avg(groq_results, "latency_s"), avg(ollama_results, "latency_s")),
        ("p50 latency (s)", percentile(groq_results, "latency_s", 50), percentile(ollama_results, "latency_s", 50)),
        ("p95 latency (s)", percentile(groq_results, "latency_s", 95), percentile(ollama_results, "latency_s", 95)),
        ("p99 latency (s)", percentile(groq_results, "latency_s", 99), percentile(ollama_results, "latency_s", 99)),
        ("Avg tokens/sec", avg(groq_results, "tokens_per_sec"), avg(ollama_results, "tokens_per_sec")),
    ]
    for row in metrics:
        ws_s.append(list(row))

    _autofit_columns(ws_g)
    _autofit_columns(ws_o)
    _autofit_columns(ws_s)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_headers(ws, headers: list) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.delete_rows(1)
    ws.insert_rows(1)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


@router.message(Command("model"))
async def cmd_model(message: Message) -> None:
    """Задача 17: /model pull <name> — загрузка модели с прогресс-баром."""
    parts = (message.text or "").split()
    if len(parts) < 3 or parts[1] != "pull":
        available = await ollama_service.list_models()
        models_text = "\n".join(f"• {m}" for m in available) if available else "Нет загруженных моделей"
        await message.answer(
            f"Использование: <code>/model pull &lt;имя&gt;</code>\n\n"
            f"Загруженные модели:\n{models_text}",
            parse_mode="HTML",
        )
        return

    model_name = parts[2]
    sent = await message.answer(f"⬇️ Начинаю загрузку модели <b>{model_name}</b>…", parse_mode="HTML")

    last_update = 0.0
    last_status = ""

    async for status in ollama_service.pull_model_with_progress(model_name):
        now = asyncio.get_event_loop().time()
        if now - last_update >= 5.0 and status != last_status:
            try:
                await sent.edit_text(
                    f"⬇️ Загрузка <b>{model_name}</b>:\n<code>{status}</code>",
                    parse_mode="HTML",
                )
                last_update = now
                last_status = status
            except Exception:
                pass

    await sent.edit_text(
        f"✅ Модель <b>{model_name}</b> успешно загружена!",
        parse_mode="HTML",
    )
